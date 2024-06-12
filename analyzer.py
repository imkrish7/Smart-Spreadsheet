import openpyxl
import pandas as pd
from spire.xls import *
from spire.xls.common import *
# from tabula import read_pdf
# from tabulate import tabulate

# from eparse.core import get_df_from_file, df_find_tables

# def read_table(file_name: str, sheet_name: str):
    # wb = openpyxl.load_workbook(file_name, read_only=False, data_only=True)
    # wb.
    # for sheetname in wb.sheetnames:
    #     print(wb[sheetname]);
    #     sheet = wb[sheetname]
    #     print(sheet.tables)
    #     # if sheet_name in sheet.tables:
    #     tbl = sheet.tables[sheetname]
    #     tbl_range = tbl.ref
    #     data = sheet[tbl_range]
    #     content = [[cell.value for cell in row] for row in data]
    #     header = content[0]
    #     res = content[1:]
    #     df = pd.DataFrame(res, columns=header)
    #     print("---print---")
    #     print(df)
    # wb.close()

# def read_table(file_name: str, sheet_name: str):
#     # tables = get_df_from_file(file_name)
#     # # print([table for table in tables])
#     # for table in tables:
#     #     for row in table:
#     #         print(row)
#     df = pd.read_excel(file_name)
#     print(df)


def read_table(file_name: str, sheet_name: str):
    wb = Workbook();
    wb.LoadFromFile(file_name)
    sheet = wb.Worksheets[0]
    wb.ConverterSetting.SheetFitToPage = True

    # Get a specific worksheet (index starts at zero)
    

    # Get the PageSetup object
    pageSetup = sheet.PageSetup

    # Set page margins
    pageSetup.TopMargin = 0.1
    pageSetup.BottomMargin = 0.1
    pageSetup.LeftMargin = 0.1
    pageSetup.RightMargin = 0.5

    # The PDF page size is adjusted to fit the content area
    wb.ConverterSetting.SheetFitToPage = True
    sheet.SaveToPdf('./js-sol/out.pdf')
    wb.Dispose()

    # df = read_pdf('out.pdf', pages='all')
    # print(tabulate(df));


read_table('tests/example_1.xlsx', 'Analysis Output');